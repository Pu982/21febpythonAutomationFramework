# =========================================================
# Excel Utilities - Read / Write Excel Operations
# =========================================================

from openpyxl import load_workbook, Workbook
import logging
import os

logger = logging.getLogger(__name__)


class ExcelUtils:
    """
    Utility class for Excel operations
    """

    # =========================================================
    # WORKBOOK METHODS
    # =========================================================

    @staticmethod
    def load_excel(file_path):
        """
        Load Excel workbook
        """
        try:
            wb = load_workbook(file_path)

            logger.info(f"Workbook loaded: {file_path}")

            return wb

        except Exception as e:
            logger.error(f"Unable to load workbook | Error: {e}")
            raise

    @staticmethod
    def create_excel(file_path):
        """
        Create new Excel workbook
        """
        try:
            wb = Workbook()

            wb.save(file_path)

            logger.info(f"Workbook created: {file_path}")

        except Exception as e:
            logger.error(f"Unable to create workbook | Error: {e}")
            raise

    @staticmethod
    def save_excel(workbook, file_path):
        """
        Save workbook
        """
        try:
            workbook.save(file_path)

            logger.info(f"Workbook saved: {file_path}")

        except Exception as e:
            logger.error(f"Unable to save workbook | Error: {e}")
            raise

    # =========================================================
    # SHEET METHODS
    # =========================================================

    @staticmethod
    def get_sheet(file_path, sheet_name="Sheet"):
        """
        Get worksheet object
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        logger.info(f"Accessed sheet: {sheet_name}")

        return ws

    @staticmethod
    def get_sheet_names(file_path):
        """
        Get all sheet names
        """
        wb = load_workbook(file_path)

        sheet_names = wb.sheetnames

        logger.info(f"Sheet names: {sheet_names}")

        return sheet_names

    @staticmethod
    def create_sheet(file_path, sheet_name):
        """
        Create new sheet
        """
        wb = load_workbook(file_path)

        wb.create_sheet(sheet_name)

        wb.save(file_path)

        logger.info(f"Sheet created: {sheet_name}")

    @staticmethod
    def delete_sheet(file_path, sheet_name):
        """
        Delete sheet
        """
        wb = load_workbook(file_path)

        del wb[sheet_name]

        wb.save(file_path)

        logger.info(f"Sheet deleted: {sheet_name}")

    # =========================================================
    # READ METHODS
    # =========================================================

    @staticmethod
    def read_excel(file_path, sheet_name="Sheet"):
        """
        Read complete sheet data
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        data = []

        for row in ws.iter_rows(values_only=True):
            data.append(row)

        logger.info(f"Excel data read from: {sheet_name}")

        return data

    @staticmethod
    def read_cell(file_path, sheet_name, row, column):
        """
        Read single cell value
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        value = ws.cell(
            row=row,
            column=column
        ).value

        logger.info(
            f"Read cell value at Row {row}, Column {column}"
        )

        return value

    @staticmethod
    def read_row(file_path, sheet_name, row):
        """
        Read complete row
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        row_data = []

        for cell in ws[row]:
            row_data.append(cell.value)

        logger.info(f"Read row: {row}")

        return row_data

    @staticmethod
    def read_column(file_path, sheet_name, column):
        """
        Read complete column
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        column_data = []

        for cell in ws.iter_cols(
            min_col=column,
            max_col=column,
            values_only=True
        ):
            column_data.extend(cell)

        logger.info(f"Read column: {column}")

        return column_data

    # =========================================================
    # WRITE METHODS
    # =========================================================

    @staticmethod
    def write_cell(
        file_path,
        sheet_name,
        row,
        column,
        value
    ):
        """
        Write value into cell
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        ws.cell(
            row=row,
            column=column
        ).value = value

        wb.save(file_path)

        logger.info(
            f"Written value at Row {row}, Column {column}"
        )

    @staticmethod
    def write_row(
        file_path,
        sheet_name,
        row,
        data
    ):
        """
        Write complete row
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        for col_num, value in enumerate(data, start=1):
            ws.cell(
                row=row,
                column=col_num
            ).value = value

        wb.save(file_path)

        logger.info(f"Written row: {row}")

    @staticmethod
    def append_row(
        file_path,
        sheet_name,
        data
    ):
        """
        Append row at end
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        ws.append(data)

        wb.save(file_path)

        logger.info("Row appended successfully")

    # =========================================================
    # ROW / COLUMN COUNT METHODS
    # =========================================================

    @staticmethod
    def get_row_count(file_path, sheet_name):
        """
        Get total number of rows
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        rows = ws.max_row

        logger.info(f"Total rows: {rows}")

        return rows

    @staticmethod
    def get_column_count(file_path, sheet_name):
        """
        Get total number of columns
        """
        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        columns = ws.max_column

        logger.info(f"Total columns: {columns}")

        return columns

    # =========================================================
    # FILE VALIDATION METHODS
    # =========================================================

    @staticmethod
    def file_exists(file_path):
        """
        Check whether file exists
        """
        exists = os.path.exists(file_path)

        logger.info(f"File exists: {exists}")

        return exists

    @staticmethod
    def sheet_exists(file_path, sheet_name):
        """
        Check whether sheet exists
        """
        wb = load_workbook(file_path)

        exists = sheet_name in wb.sheetnames

        logger.info(f"Sheet exists: {exists}")

        return exists